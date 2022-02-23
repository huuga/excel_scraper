from openpyxl import load_workbook
import pandas as pd

def names_case_normalizer(winners):
	for winner in winners:
		person = winner.split()
		new_person = ''
		for person_ in person:
			new_person += person_.capitalize() + ' ' 
		winners[winners.index(winner)] = new_person.strip()


def list_stripper(lst):
	for i in range(len(lst)):
		lst[i] = lst[i].strip()
	return lst


def winners_list():
	try:
		wb = load_workbook('./Медальный зачет X ОРЧ (2).xlsx')
		sheet = wb['Лист1']
		winners = []

		for cells in sheet['E1':'E' + str(sheet.max_row)]:
			if cells[0].value == 1 and sheet['B' + str(cells[0].row)].value.strip() == 'Основная':
				winners.append(sheet['D' + str(cells[0].row)].value)

		for winner in winners:
			if '/' in winner:
				i = winners.index(winner)
				winners = winners[:i] + list_stripper(winner.split('/')) + winners[i + 1:]
		wb.close	
		return winners
	except Exception as e:
		print(e)


def is_winner_in_list(winner, sheet):
	participants_list = []

	for x in sheet['E1':'E' + str(sheet.max_row)]:
		if x[0].value:
			participants_list.append(x[0].value.strip().lower())

	return winner.lower() in participants_list

def winners_data_extract(winners):
	try:
		wb = load_workbook('./!!!СВОД ЗАЯВОК (15.12.2021) (1).xlsx')
		sheet = wb['Заявка участников и экспертов']
		result = []

		for winner in winners:
			if is_winner_in_list(winner, sheet):
				for x in sheet['E1':'E' + str(sheet.max_row)]:
					if x[0].value:
						if winner.casefold() == x[0].value.strip().casefold():
							r = str(x[0].row)
							user_info = [
								sheet['B' + r].value,
								sheet['C' + r].value,
								x[0].value,
								sheet['F' + r].value.strftime("%d.%m.%Y"),
								sheet['I' + r].value,
								sheet['L' + r].value,
								sheet['M' + r].value,
								sheet['N' + r].value,
								sheet['O' + r].value
							]
							result.append(user_info)
			else:
				result.append(['', '', winner, '', '', '', '', '', ''])
		wb.close
		return result
	except Exception as e:
		print(e)



if __name__ == "__main__":
	winners = winners_list()
	names_case_normalizer(winners)
	total_list_of_winners = winners_data_extract(winners)
	df = pd.DataFrame(total_list_of_winners)
	writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
	df.to_excel(writer, sheet_name='Сборная Якутии', index=False)
	writer.save()
	
